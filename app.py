import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="Color Clarity Automation")

st.title("Color Clarity Automation")

# =========================================
# FILE UPLOADERS
# =========================================

color_clarity_file = st.file_uploader(
    "Upload Color Clarity File",
    type=["xlsx"]
)

pending_video_file = st.file_uploader(
    "Upload Pending Video File",
    type=["xlsx"]
)

certify_file = st.file_uploader(
    "Upload Certify Color Stone File",
    type=["xlsx"]
)

# =========================================
# SIZE GROUP FUNCTION
# =========================================

def get_size_group(value):

    try:

        value = float(value)

        if 1.00 <= value <= 1.49:
            return "1.00-1.49"

        elif 1.50 <= value <= 1.99:
            return "1.50-1.99"

        elif 2.00 <= value <= 2.99:
            return "2.00-2.99"

        elif 3.00 <= value <= 3.99:
            return "3.00-3.99"

        elif 4.00 <= value <= 4.99:
            return "4.00-4.99"

        elif 5.00 <= value <= 5.99:
            return "5.00-5.99"

        elif 6.00 <= value <= 6.99:
            return "6.00-6.99"

        elif 7.00 <= value <= 7.99:
            return "7.00-7.99"

        elif 8.00 <= value <= 8.99:
            return "8.00-8.99"

        elif 9.00 <= value <= 9.99:
            return "9.00-9.99"

        elif 10.00 <= value <= 10.99:
            return "10.00-10.99"

        else:
            return "Out of Range"

    except:
        return "Out of Range"

# =========================================
# MAIN PROCESS
# =========================================

if color_clarity_file and pending_video_file and certify_file:

    # =========================================
    # LOAD FILES
    # =========================================

    df1 = pd.read_excel(color_clarity_file)

    df2 = pd.read_excel(
        pending_video_file,
        sheet_name=0
    )

    # CLEAN COLUMN NAMES
    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    # =========================================
    # STEP 1 - KEEP ONLY IGI
    # =========================================

    df1 = df1[
        df1["Lab"]
        .astype(str)
        .str.upper()
        == "IGI"
    ]

    # =========================================
    # STEP 2 - COLOR CLEANING
    # =========================================

    allowed_colors = [
        "BLUE",
        "BROWN",
        "YELLOW",
        "GREEN",
        "ORANGE",
        "PINK"
    ]

    def clean_color(color):

        if pd.isna(color):
            return None

        color_upper = str(color).upper()

        for allowed in allowed_colors:

            if allowed in color_upper:
                return allowed

        return None

    df1["Color"] = df1["Color"].apply(clean_color)

    df1 = df1[
        df1["Color"].notna()
    ]

    # =========================================
    # STEP 3 - KEEP REQUIRED COLUMNS
    # =========================================

    required_columns = [
        "Serial #",
        "Shape",
        "Color",
        "Grade",
        "Cts.",
        "Lab",
        "H&&A",
        "Certificate #"
    ]

    df1 = df1[required_columns]

    # =========================================
    # STEP 4 - SHAPE CLEANING
    # =========================================

    df1["Shape"] = (
        df1["Shape"]
        .astype(str)
        .str.upper()
    )

    df1["Shape"] = df1["Shape"].replace({
        "CUSHION MODIFIED": "CUSHION",
        "CUSHION BRILLIANT": "CUSHION"
    })

    # =========================================
    # STEP 5 - REMOVE PREFIXES
    # =========================================

    remove_prefixes = (
        "NY",
        "V",
        "DM",
        "DC"
    )

    df1 = df1[
        ~df1["Serial #"]
        .astype(str)
        .str.upper()
        .str.startswith(remove_prefixes)
    ]

    # =========================================
    # STEP 6 - STATUS LOGIC
    # =========================================

    customer_values = [
        "GOODS IN TRANSIT FROM OVERSEAS",
        "GOODS IN OFFICE - PARCEL PAPERS BEING MADE",
        "JCK 2026"
    ]

    mask = (
        df2["Customer"]
        .astype(str)
        .str.upper()
        .isin(customer_values)
    )

    df2.loc[
        mask &
        (
            df2["Status"]
            .astype(str)
            .str.upper() == "ONMEMO"
        ),
        "Status"
    ] = "Inhand"

    # =========================================
    # STEP 7 - MATCH STATUS
    # =========================================

    status_map = (
        df2
        .set_index("Lot #")["Status"]
        .to_dict()
    )

    serial_index = df1.columns.get_loc("Serial #")

    df1.insert(
        serial_index + 1,
        "Status",
        df1["Serial #"].map(status_map)
    )

    # =========================================
    # STEP 8 - SIZE GROUP
    # =========================================

    cts_index = df1.columns.get_loc("Cts.")

    df1.insert(
        cts_index + 1,
        "SIZE GROUP",
        df1["Cts."].apply(get_size_group)
    )

    # REMOVE OUT OF RANGE
    df1 = df1[
        df1["SIZE GROUP"] != "Out of Range"
    ]

    # =========================================
    # STEP 9 - CREATE PIVOT TABLE
    # =========================================

    pivot_df = df1[
        df1["Status"]
        .astype(str)
        .str.upper() == "INHAND"
    ]

    pivot_table = pd.pivot_table(
        pivot_df,
        index=["Shape", "SIZE GROUP"],
        columns="Color",
        values="Serial #",
        aggfunc="count",
        fill_value=0
    )

    pivot_table = pivot_table.reset_index()

    # =========================================
    # STEP 10 - LOAD CERTIFY FILE
    # =========================================

    wb = load_workbook(certify_file)

    color_sheets = [
        "BLUE",
        "BROWN",
        "GREEN",
        "ORANGE",
        "PINK",
        "YELLOW"
    ]

    possible_shapes = [
        "HEART",
        "OVAL",
        "CUSHION",
        "PEAR",
        "MARQUISE",
        "PRINCESS",
        "RADIANT",
        "EMERALD",
        "ASSCHER",
        "BAGUETTE",
        "ANGEL",
        "BUTTERFLY",
        "FLOWER",
        "LONG RADIANT"
    ]

    # =========================================
    # STEP 11 - PROCESS COLOR SHEETS
    # =========================================

    for color_name in color_sheets:

        if color_name not in wb.sheetnames:
            continue

        ws = wb[color_name]

        # =========================================
        # FIND SHAPE TABLES
        # =========================================

        for row in range(1, ws.max_row + 1):

            for col in range(1, ws.max_column + 1):

                cell_value = ws.cell(
                    row=row,
                    column=col
                ).value

                if not cell_value:
                    continue

                value_upper = (
                    str(cell_value)
                    .strip()
                    .upper()
                )

                # =========================================
                # DETECT SHAPE
                # =========================================

                if value_upper in possible_shapes:

                    current_shape = value_upper

                    header_row = row + 1
                    data_start_row = row + 2

                    size_col = None
                    inhand_col = None

                    # =========================================
                    # FIND HEADERS
                    # =========================================

                    for search_col in range(1, ws.max_column + 1):

                        header_value = ws.cell(
                            row=header_row,
                            column=search_col
                        ).value

                        if not header_value:
                            continue

                        header_upper = (
                            str(header_value)
                            .strip()
                            .upper()
                        )

                        if header_upper == "SIZE":
                            size_col = search_col

                        elif header_upper == "INHAND":
                            inhand_col = search_col

                    # =========================================
                    # PROCESS TABLE ROWS
                    # =========================================

                    current_data_row = data_start_row

                    while True:

                        size_value = ws.cell(
                            row=current_data_row,
                            column=size_col
                        ).value

                        # STOP AT TOTAL
                        if str(size_value).strip().upper() == "TOTAL":
                            break

                        try:

                            size_float = float(size_value)

                            size_group = get_size_group(
                                size_float
                            )

                            # =========================================
                            # MATCH PIVOT
                            # =========================================

                            match_df = pivot_table[
                                (
                                    pivot_table["Shape"]
                                    .astype(str)
                                    .str.upper()
                                    == current_shape
                                )
                                &
                                (
                                    pivot_table["SIZE GROUP"]
                                    == size_group
                                )
                            ]

                            if not match_df.empty:

                                if color_name in match_df.columns:

                                    pivot_value = (
                                        match_df.iloc[0][color_name]
                                    )

                                    if pd.isna(pivot_value):
                                        pivot_value = 0

                                    # =========================================
                                    # WRITE INHAND VALUE
                                    # =========================================

                                    ws.cell(
                                        row=current_data_row,
                                        column=inhand_col
                                    ).value = int(pivot_value)

                        except:
                            pass

                        current_data_row += 1

    # =========================================
    # SAVE FINAL OUTPUT
    # =========================================

    final_output = "Final_Output.xlsx"

    with pd.ExcelWriter(
        final_output,
        engine="openpyxl"
    ) as writer:

        df1.to_excel(
            writer,
            sheet_name="Sheet1",
            index=False
        )

        pivot_table.to_excel(
            writer,
            sheet_name="Sheet2",
            index=False
        )

    # =========================================
    # SAVE UPDATED CERTIFY FILE
    # =========================================

    certify_output = "Updated_Certify_Color_Stone.xlsx"

    wb.save(certify_output)

    st.success("Automation Completed Successfully!")

    # =========================================
    # DOWNLOAD BUTTONS
    # =========================================

    with open(final_output, "rb") as file1:

        st.download_button(
            label="Download Final Output",
            data=file1,
            file_name=final_output,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with open(certify_output, "rb") as file2:

        st.download_button(
            label="Download Updated Certify File",
            data=file2,
            file_name=certify_output,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
